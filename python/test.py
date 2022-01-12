import os
import re
from openpyxl import Workbook
# import sys

def write_file():
    rootdir = './log'
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    with open("log.log", "a+") as f:
        for i in range(0, len(list)):
            path = os.path.join(rootdir, list[i])
            if os.path.isfile(path):
                with open(path, 'r') as p:
                    f.writelines(p.readlines())

def split_actor():
    list = []
    str = ''
    is_actor_re = re.compile(r'del_data_offline\sActorId')
    with open('log.log', 'r') as f:
        line = f.readline()
        while(line != ''):
            if re.search(is_actor_re, line) != None:
                if str != '': list.append(str)
                str = line
            else:
                str += line
            line = f.readline()
    # print(list[0])
    return list

# 处理文件中的字符串
def pattern_str():
    are = re.compile(r'del_data_offline\sActorId\s(\d+)')
    mre = re.compile(r'del_mail_list\sSucc\sRet\s(\d+)')
    rre = re.compile(r'reduce_item_role\sRemainList\s(\[.*?\]\s=>\s\[.*?\])')
    ire = re.compile(r'reduce_item_db\sSucc.*?Uid\s(\d+).*?AddCount\s(-\d+)')
    # sys.stdout = open('output.txt', 'wt')

    wb = Workbook() # create a excel
    ws = wb.active
    ws.title = 'PetSoul Stone Data'
    ws.append(['ActorId', 'MailNum', 'RoleData', 'ItemData'])
    row = 2 # 从第二行开始填数据

    for actor in split_actor()[:1]:
        actorid = int(re.findall(are, actor)[0])

        mailret = re.findall(mre, actor)
        mailnum = 0 if mailret == [] else int(mailret[0])

        roleret = re.findall(rre, actor)
        roledata = '[] => []' if roleret == [] else roleret[0]
        # if roleret == []:
        #     roledata = []
        # else:
        #     roledict = {}
        #     From, To = eval(roleret[0][0]), eval(roleret[0][1])
        # print(roledata)

        itemlist0 = re.findall(ire, actor)
        itemlist1 = []
        for (id, cnt) in itemlist0:
            itemlist1.append({int(id), int(cnt)})
        itemlist = str(itemlist1)

        # 将匹配到的数据写入表格
        ws.cell(row=row, column=1, value=str(actorid))
        ws.cell(row=row, column=2, value=mailnum)
        ws.cell(row=row, column=3, value=roledata)
        ws.cell(row=row, column=4, value=itemlist)
        print('ActorId:', actorid, 'Mail:', mailnum, roledata, itemlist)

    # 保存文件
    wb.save(filename = 'out.xlsx')

if __name__ == '__main__':
    # write_file()
    pattern_str()