import re
from openpyxl import *
from openpyxl.utils import *

def get_item(itemstr):
    # 正则匹配字符串中的数字
    pa = re.compile(r'(\d+\*\d+)')
    if itemstr == None: return []
    list = []
    for s in re.findall(pa, itemstr):
        nums = s.split('*')
        item = '(' + nums[0] + ', ' + nums[1] + ')'
        # 通过 eval 函数将字符串转成一个结构
        list.append(eval(item))
    return list

# sys.stdout 可以将输出重定向到文件
# sys.stdout = open('output.log', 'wt')
filename = 'new.xlsx'                                   # filename
wb = load_workbook(filename=filename)                   # 用 openpyxl 加载一个Excel文件
ws = wb['1']                                            # sheetname 取一个表

for row in range(2, 206):                               # from row1 to row2
    rowstr = str(row)
    sid = ws[get_column_letter(1) + rowstr].value       # sid col
    aid = ws[get_column_letter(2) + rowstr].value       # aid col
    item1 = ws[get_column_letter(3) + rowstr].value     # item1 col
    item2 = ws[get_column_letter(4) + rowstr].value     # item2 col
    itemlist = get_item(item1) + get_item(item2)
    line = '{' + str(sid) + ', ' + str(aid) + ', ' + str(itemlist) + '},'
    # 转换格式
    print(line.replace('(', '{').replace(')', '}'))